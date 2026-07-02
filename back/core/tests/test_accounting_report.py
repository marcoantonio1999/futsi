from io import BytesIO

import pytest
from django.db import connection
from django.test.utils import CaptureQueriesContext
from openpyxl import load_workbook

from core.models import DiscountStatus, ExpenseStatus
from core.tests.factories import (
    make_attendance_record,
    make_attendance_session,
    make_charge,
    make_discount,
    make_expense,
    make_invoice,
    make_payment,
    make_site,
    make_student,
    make_user,
)


pytestmark = [pytest.mark.api, pytest.mark.django_db]


def test_accounting_export_returns_valid_xlsx_for_accounting(auth_client):
    site = make_site(name="QA Accounting", code="qa-accounting")
    accounting = make_user(role="accounting", primary_site=site)
    cashier = make_user(role="cashier", primary_site=site)
    student = make_student(site=site)
    charge = make_charge(student=student, site=site, created_by=accounting)
    payment = make_payment(charge, received_by=cashier)
    make_discount(charge=charge, status=DiscountStatus.APPROVED, requested_by=cashier, approved_by=accounting)
    make_expense(site=site, captured_by=cashier, approved_by=accounting, status=ExpenseStatus.APPROVED)
    session = make_attendance_session(site=site, captured_by=cashier)
    make_attendance_record(session=session, student=student, captured_by=cashier, had_debt_at_capture=True)
    make_invoice(site=site, student=student, guardian=student.guardian, charge=charge, payment=payment, issued_by=accounting)
    client, _payload, _user = auth_client(user=accounting)

    with CaptureQueriesContext(connection) as captured:
        response = client.get("/api/reports/accounting.xlsx")

    assert response.status_code == 200
    assert response["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert response.content[:2] == b"PK"
    assert len(captured) <= 8
    invoice_query = next(query["sql"] for query in captured if 'FROM "invoices"' in query["sql"])
    assert "xml_content" not in invoice_query

    workbook = load_workbook(BytesIO(response.content), read_only=True)
    assert {
        "Resumen",
        "Pagos",
        "Cargos",
        "Gastos",
        "Descuentos",
        "Asistencia con adeudo",
        "Facturas",
    }.issubset(set(workbook.sheetnames))


def test_accounting_export_is_forbidden_for_cashier(auth_client):
    site = make_site(name="QA Accounting Forbidden", code="qa-accounting-forbidden")
    cashier = make_user(role="cashier", primary_site=site)
    client, _payload, _user = auth_client(user=cashier)

    response = client.get("/api/reports/accounting.xlsx")

    assert response.status_code == 403
