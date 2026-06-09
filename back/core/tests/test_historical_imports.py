from io import BytesIO

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook

from core.models import HistoricalImport, Payment, Expense


pytestmark = [pytest.mark.api, pytest.mark.django_db]


def build_historical_excel():
    workbook = Workbook()
    workbook.remove(workbook.active)

    income = workbook.create_sheet("INGRESOS SEDES")
    income.append(["Mayo", "", "", "ACA BOSQUES", "Total"])
    income.append(["", "", "", "", ""])
    income.append(["", "MENS", "Mensualidad historica", 1500, 1500])

    expense = workbook.create_sheet("GASTOS SEDES")
    expense.append(["Mayo", "", "", "LGA BOSQUES", "Total"])
    expense.append(["", "", "", "", ""])
    expense.append(["", "ARB", "Arbitraje historico", 300, 300])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def build_discrepancy_excel():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Lista Verificacion Adeudos"
    worksheet.cell(row=1, column=1, value="Roma")
    worksheet.cell(row=7, column=4, value=1500)
    worksheet.cell(row=10, column=7, value="ENERO")
    worksheet.cell(row=11, column=1, value="No.")
    worksheet.cell(row=11, column=2, value="Categoria")
    worksheet.cell(row=11, column=3, value="Nombre")
    worksheet.cell(row=11, column=4, value="Padre o Tutor")
    worksheet.cell(row=11, column=5, value="Telefono")
    worksheet.cell(row=11, column=7, value="Clases")
    worksheet.cell(row=11, column=8, value="Folio")
    worksheet.cell(row=11, column=9, value="$")

    worksheet.cell(row=13, column=1, value="SIN PAGO")
    worksheet.cell(row=14, column=1, value=1)
    worksheet.cell(row=14, column=2, value="Sub-10")
    worksheet.cell(row=14, column=3, value="Alumno Sin Pago")
    worksheet.cell(row=14, column=4, value="Tutor Sin Pago")
    worksheet.cell(row=14, column=5, value="5551112222")
    worksheet.cell(row=14, column=7, value=4)

    worksheet.cell(row=21, column=1, value="CON PAGO INCOMPLETO")
    worksheet.cell(row=22, column=1, value=2)
    worksheet.cell(row=22, column=2, value="Sub-12")
    worksheet.cell(row=22, column=3, value="Alumno Parcial")
    worksheet.cell(row=22, column=4, value="Tutor Parcial")
    worksheet.cell(row=22, column=5, value="5553334444")
    worksheet.cell(row=22, column=7, value=3)
    worksheet.cell(row=22, column=8, value="FOL-123")
    worksheet.cell(row=22, column=9, value=500)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def test_admin_can_preview_and_commit_historical_excel(admin_client):
    upload = SimpleUploadedFile(
        "historico-test.xlsx",
        build_historical_excel(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    preview_response = admin_client.post(
        "/api/historical-imports/preview/",
        {"file": upload, "notes": "QA Sprint 4"},
        format="multipart",
    )

    assert preview_response.status_code == 201, preview_response.content
    body = preview_response.json()
    assert body["status"] == "draft"
    assert body["row_count"] == 2
    assert {row["row_type"] for row in body["rows"]} == {"income", "expense"}

    commit_response = admin_client.post(
        f"/api/historical-imports/{body['id']}/commit/",
        {
            "signature_name": "QA Admin",
            "signature_role": "Automated Test",
            "rows": [
                {
                    "id": row["id"],
                    "row_type": row["row_type"],
                    "site": row["site"],
                    "concept": row["concept"],
                    "amount": row["amount"],
                    "record_date": row["record_date"],
                    "skip": False,
                }
                for row in body["rows"]
            ],
        },
        format="json",
    )

    assert commit_response.status_code == 200, commit_response.content
    assert commit_response.json()["status"] == "committed"
    assert Payment.objects.filter(notes__contains="Historico Excel cerrado").exists()
    assert Expense.objects.filter(description__contains="Arbitraje historico").exists()


def test_cashier_cannot_upload_historical_excel(login_client):
    client, _user = login_client("caja.roma", "demo12345")
    upload = SimpleUploadedFile(
        "historico-test.xlsx",
        build_historical_excel(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    response = client.post("/api/historical-imports/preview/", {"file": upload}, format="multipart")

    assert response.status_code == 403
    assert HistoricalImport.objects.count() == 0


def test_commit_requires_signature(admin_client):
    upload = SimpleUploadedFile(
        "historico-test.xlsx",
        build_historical_excel(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    preview_response = admin_client.post("/api/historical-imports/preview/", {"file": upload}, format="multipart")

    response = admin_client.post(
        f"/api/historical-imports/{preview_response.json()['id']}/commit/",
        {"signature_name": "", "rows": []},
        format="json",
    )

    assert response.status_code == 400


def test_historical_discrepancies_are_detected_by_site(admin_client):
    upload = SimpleUploadedFile(
        "adeudos-historicos.xlsx",
        build_discrepancy_excel(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    preview_response = admin_client.post(
        "/api/historical-imports/preview/",
        {"file": upload, "notes": "Adeudos historicos"},
        format="multipart",
    )

    assert preview_response.status_code == 201, preview_response.content
    body = preview_response.json()
    assert body["summary"]["discrepancies"] == 2
    assert {row["row_type"] for row in body["rows"]} == {"discrepancy"}

    commit_response = admin_client.post(
        f"/api/historical-imports/{body['id']}/commit/",
        {
            "signature_name": "QA Admin",
            "signature_role": "Contador",
            "rows": [
                {
                    "id": row["id"],
                    "row_type": row["row_type"],
                    "site": row["site"],
                    "concept": row["concept"],
                    "amount": row["amount"],
                    "record_date": row["record_date"],
                    "skip": False,
                }
                for row in body["rows"]
            ],
        },
        format="json",
    )

    assert commit_response.status_code == 200, commit_response.content

    report_response = admin_client.get("/api/historical-imports/discrepancies/")
    assert report_response.status_code == 200, report_response.content
    report = report_response.json()
    assert report["totals"]["historical_cases"] == 2
    assert report["totals"]["high_risk"] == 1
    assert {item["discrepancy_type"] for item in report["items"]} == {"no_payment_no_folio", "partial_payment"}
    assert report["summary"][0]["site_name"] == "Roma"
