from __future__ import annotations

import re
import tempfile
import unicodedata
from pathlib import Path
from urllib.error import HTTPError
from urllib.parse import quote
from urllib.request import Request, urlopen

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from core.models import Player, Site, Team, Tournament
from core.services.supabase_storage import service_role_key, storage_uri, supabase_url, upload_private_file


SKIPPED_SHEETS = {"BASE", "BRASILEÑA"}
DEFAULT_SITE_ID = 27
DEFAULT_TOURNAMENT_NAME = "BRASILEÑA"
NOISE_WORDS = {
    "arbitraje",
    "baja",
    "sancionado",
    "guantes",
    "brasileña",
    "brasilena",
}


def normalize_text(value: str) -> str:
    text = unicodedata.normalize("NFKD", value or "")
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text).strip()


def normalize_key(value: str) -> str:
    return normalize_text(value).casefold()


def safe_path_part(value: str) -> str:
    text = unicodedata.normalize("NFKD", value or "")
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^a-zA-Z0-9._-]+", "_", text).strip("_")
    return text or "player"


def is_person_name(value) -> bool:
    if not isinstance(value, str):
        return False
    text = normalize_text(value)
    lowered = text.casefold()
    if len(text) < 3 or lowered in {"x", "o", ".", "0"}:
        return False
    if any(word in lowered for word in NOISE_WORDS):
        return False
    letters = re.findall(r"[A-Za-zÁÉÍÓÚÜÑáéíóúüñ]+", text)
    return len("".join(letters)) >= 3


def ensure_private_bucket(bucket: str) -> None:
    endpoint = f"{supabase_url()}/storage/v1/bucket"
    payload = f'{{"id":"{bucket}","name":"{bucket}","public":false}}'.encode("utf-8")
    key = service_role_key()
    request = Request(
        endpoint,
        data=payload,
        method="POST",
        headers={
            "Authorization": f"Bearer {key}",
            "apikey": key,
            "Content-Type": "application/json",
        },
    )
    try:
        with urlopen(request, timeout=60) as response:
            if response.status >= 400:
                raise RuntimeError(f"Supabase Storage respondio {response.status} al crear bucket.")
    except HTTPError as exc:
        body = exc.read().decode("utf-8", errors="ignore")
        if exc.code == 400 and ("already exists" in body.lower() or "already_exist" in body.lower()):
            return
        if exc.code == 409:
            return
        raise RuntimeError(f"No se pudo crear/verificar bucket privado {bucket}: HTTP {exc.code}: {body}") from exc


def closest_name_for_image(ws, image) -> tuple[str, int, int] | None:
    anchor = image.anchor._from
    image_row = anchor.row + 1
    image_col = anchor.col + 1
    candidates: list[tuple[int, str, int, int]] = []
    for row in range(max(1, image_row - 2), min(ws.max_row, image_row + 2) + 1):
        for col in range(max(1, image_col - 1), min(ws.max_column, image_col + 4) + 1):
            value = ws.cell(row, col).value
            if not is_person_name(value):
                continue
            text = normalize_text(str(value))
            # Photos are usually anchored 1-2 columns left of the player name.
            distance = abs(row - image_row) * 10 + abs(col - (image_col + 1))
            candidates.append((distance, text, row, col))
    if not candidates:
        return None
    _, name, row, col = sorted(candidates, key=lambda item: item[0])[0]
    return name, row, col


def image_extension(image) -> str:
    fmt = (getattr(image, "format", "") or "").lower()
    if fmt in {"jpeg", "jpg"}:
        return ".jpg"
    if fmt == "png":
        return ".png"
    return ".jpg"


def write_image_temp(image, name: str) -> Path:
    suffix = image_extension(image)
    handle = tempfile.NamedTemporaryFile(suffix=suffix, delete=False)
    handle.write(image._data())
    handle.close()
    return Path(handle.name)


class Command(BaseCommand):
    help = "Importa jugadores adultos y fotos incrustadas desde un Excel a Player.photo_url con Supabase Storage privado."

    def add_arguments(self, parser):
        parser.add_argument("--xlsx", required=True)
        parser.add_argument("--site-id", type=int, default=DEFAULT_SITE_ID)
        parser.add_argument("--tournament-name", default=DEFAULT_TOURNAMENT_NAME)
        parser.add_argument("--bucket", default="adult-private-photos")
        parser.add_argument("--team-name-source", choices=["sheet", "header"], default="sheet")
        parser.add_argument("--apply", action="store_true")

    def handle(self, *args, **options):
        xlsx = Path(options["xlsx"])
        if not xlsx.exists():
            raise CommandError(f"No existe el Excel: {xlsx}")

        site = Site.objects.get(id=options["site_id"])
        tournament_name = options["tournament_name"].strip()
        bucket = options["bucket"].strip()
        apply_changes = options["apply"]
        team_name_source = options["team_name_source"]

        workbook = load_workbook(xlsx, data_only=True)
        imported_rows = []
        skipped = []

        for ws in workbook.worksheets:
            if ws.title.strip().upper() in SKIPPED_SHEETS:
                continue
            header_team_name = normalize_text(str(ws.cell(1, 4).value or ws.cell(2, 4).value or ws.title))
            team_name = normalize_text(ws.title if team_name_source == "sheet" else header_team_name)
            if not team_name:
                team_name = normalize_text(ws.title)
            seen_names = set()
            for image in getattr(ws, "_images", []):
                match = closest_name_for_image(ws, image)
                if not match:
                    skipped.append(f"{ws.title}: foto sin nombre cercano")
                    continue
                full_name, row, col = match
                name_key = normalize_key(full_name)
                if name_key in seen_names:
                    skipped.append(f"{ws.title}: {full_name} duplicado en hoja")
                    continue
                seen_names.add(name_key)
                imported_rows.append(
                    {
                        "sheet": ws.title,
                        "team_name": team_name,
                        "full_name": full_name,
                        "row": row,
                        "col": col,
                        "image": image,
                    }
                )

        created_teams = 0
        created_players = 0
        updated_players = 0
        uploaded = 0

        if apply_changes:
            ensure_private_bucket(bucket)

        with transaction.atomic():
            tournament = None
            if apply_changes:
                tournament, _ = Tournament.objects.get_or_create(
                    site=site,
                    name=tournament_name,
                    defaults={
                        "billing_type": "weekly_match",
                        "is_active": True,
                        "expected_weeks": 12,
                    },
                )

            teams_by_name = {
                normalize_key(team.name): team
                for team in Team.objects.filter(tournament__site=site, tournament__name=tournament_name)
            }

            for row in imported_rows:
                if apply_changes:
                    team_key = normalize_key(row["team_name"])
                    team = teams_by_name.get(team_key)
                    if not team:
                        team = Team.objects.create(
                            tournament=tournament,
                            name=row["team_name"],
                            representative_name=f"Capitan {row['team_name']}",
                            representative_phone="",
                            representative_email="",
                            is_active=True,
                        )
                        teams_by_name[team_key] = team
                        created_teams += 1

                    player, created = Player.objects.get_or_create(
                        team=team,
                        full_name=row["full_name"],
                        defaults={"is_active": True},
                    )
                    if created:
                        created_players += 1
                    else:
                        updated_players += 1
                        if not player.is_active:
                            player.is_active = True

                    image_path = write_image_temp(row["image"], row["full_name"])
                    object_path = (
                        f"players/{team.id}/{player.id}/"
                        f"{safe_path_part(row['full_name'])}_{row['row']}_{row['col']}{image_path.suffix.lower()}"
                    )
                    private_uri = upload_private_file(bucket, object_path, image_path, upsert=True)
                    uploaded += 1
                    player.photo_url = private_uri
                    player.save(update_fields=["photo_url", "is_active", "updated_at"])
                else:
                    team_key = normalize_key(row["team_name"])
                    if team_key not in teams_by_name:
                        created_teams += 1
                        teams_by_name[team_key] = None
                    created_players += 1

        team_counts = {}
        for row in imported_rows:
            team_counts[row["team_name"]] = team_counts.get(row["team_name"], 0) + 1

        self.stdout.write(
            self.style.SUCCESS(
                f"{'Aplicado' if apply_changes else 'Dry-run'}: {len(imported_rows)} jugadores con foto detectados, "
                f"{created_teams} equipos nuevos, {created_players} jugadores creados/detectados, "
                f"{updated_players} actualizados, {uploaded} fotos subidas, {len(skipped)} omitidos."
            )
        )
        for team_name, count in sorted(team_counts.items()):
            self.stdout.write(f"{team_name}: {count}")
        for item in skipped[:30]:
            self.stdout.write(f"omitido: {item}")
        if len(skipped) > 30:
            self.stdout.write(f"... {len(skipped) - 30} omitidos mas")
